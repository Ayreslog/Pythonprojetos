#A PRINCIPIO ESTE É CODIGO FINAl
import pyautogui
import time
import keyboard
from openpyxl import load_workbook
import pytesseract
import pyperclip
import openpyxl
import os
import re
from datetime import datetime, timedelta
from PIL import ImageGrab
import pandas as pd


# Carregar a planilha
file_path_agosto = "G:\\Drives compartilhados\\Auditoria\\LMC\\Tiburcio.xlsx"
sheet_name_agosto = "Agosto"

# Ler a planilha
df_agosto = pd.read_excel(file_path_agosto, sheet_name=sheet_name_agosto)

# Configuração do Tesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Função para clicar em uma posição
def click_position(x, y, double_click=False, right_click=False):
    if double_click:
        pyautogui.doubleClick(x, y)
    elif right_click:
        pyautogui.click(x, y, button='right')
    else:
        pyautogui.click(x, y)
    time.sleep(2)

# Função para copiar o texto da tela
def copy_text():
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(1)
    return pyperclip.paste()

# Função para verificar se a data é menor ou igual a dois dias atrás
def is_date_recent(date_str):
    try:
        date_format = "%d/%m/%Y"
        copied_date = datetime.strptime(date_str, date_format)
        two_days_ago = datetime.now() - timedelta(days=1)
        if copied_date >= two_days_ago:
            # Data não é recente; clique na posição (921, 94)
            pyautogui.click(921, 94)
            return False
        return True
    except ValueError:
        # Data inválida; clique na posição (921, 94)
        pyautogui.click(921, 94)
        return False

# Função para capturar uma área da tela e usar OCR para extrair o texto
def extract_text_from_region(left, top, width, height, psm=6):
    bbox = (left, top, left + width, top + height)
    screenshot = ImageGrab.grab(bbox)
    screenshot.save('screenshot.png')  # Salva a imagem para depuração
    text = pytesseract.image_to_string(screenshot, config=f'--psm {psm}')
    return text.strip()

# Ajuste o nome das colunas conforme necessário
col_impedimento = 'Impedimento'
col_loja_numero = 'UN'

# Função para parar o loop manualmente
def check_stop():
    if keyboard.is_pressed('esc'):
        print("Processo interrompido pelo usuário.")
        return True
    return False
while True:
# Abrir a planilha com openpyxl para garantir a integridade
 wb_agosto = load_workbook(file_path_agosto)
 ws_agosto = wb_agosto[sheet_name_agosto]

# Função principal para automatizar a extração
 def automate_extraction():
    # Definir o primeiro dia do mês
    first_day_of_month = datetime.now().replace(day=2).strftime("%d/%m/%Y")

    # Lista de posições dos produtos (todos os combustíveis da lista)
    product_positions = [
        (644, 163),  # Primeiro combustível
        (527, 181),  # Segundo combustível
        (545, 199),  # Terceiro combustível
        (547, 220),  # Quarto combustível
        #(536, 237),  # Quinto combustível
        #(524, 257),  # Sexto combustível
        #(530, 276)   # Sétimo combustível
    ]
    
    for position in product_positions:
        click_position(*position)

        # Copiar informações da tela
        text = copy_text()
        
        # Dividir o texto em linhas e colunas, ignorando a primeira linha
        rows = text.split('\n')[1:]  # Ignora a primeira linha
        data = []
        for row in rows:
            columns = row.split('\t')  # ou use row.split() para dividir por espaços
            data.append(columns)
    
        # Criar um DataFrame do pandas com os novos dados
        new_df = pd.DataFrame(data)
    
        file_path_lmc = 'LMC_Automatico.xlsx'
        #file_path_lmc = 'G:\\Drives compartilhados\\Auditoria\\LMC\\LMC_Automatico.xlsx'
    
        if os.path.exists(file_path_lmc):
            # Carregar o arquivo existente
            book = openpyxl.load_workbook(file_path_lmc)
            sheet = book.active
        
            # Encontrar a última linha preenchida
            startrow = sheet.max_row
        
            # Escrever os novos dados a partir da última linha preenchida
            for i, row in new_df.iterrows():
                for j, value in enumerate(row):
                    sheet.cell(row=startrow + i + 1, column=j + 1, value=value)
        
            # Salvar o arquivo
            book.save(file_path_lmc)
        else:
            # Se o arquivo não existe, criar um novo com os novos dados
            new_df.to_excel(file_path_lmc, index=False, header=True)
        
        # Novo passo: clicar em "Inserir item"
        click_position(208, 371)  # Inserir item
        
        # Dar um duplo clique na posição "Data LMC"
        click_position(855, 147, double_click=True)  # Data LMC
        
        # Capturar e verificar o texto da região onde a data deve estar
        date_text = extract_text_from_region(830, 135, 100, 20)  # Ajuste as coordenadas e o tamanho conforme necessário
        print(f"Texto extraído da região da data: '{date_text}'")
        
        if not date_text or date_text == 'Jata:| |':
            # Se não houver texto, digitar o primeiro dia do mês atual
            click_position(855, 147)  # Clicar na posição para focar o campo (Data LMC)
            pyautogui.typewrite(first_day_of_month)  # Digitar a data manualmente
            time.sleep(2)
            date_str = first_day_of_month  # Usar a data do primeiro dia do mês como data copiada
        else:
            # Se houver texto, clicar com botão direito e depois com botão esquerdo na posição "Copiar data"
            click_position(855, 147, right_click=True)  # Data LMC
            click_position(899, 207)  # Copiar data
            date_str = copy_text()
        
        print(f"Data copiada: '{date_str}'")
        
        # Carregar novamente o arquivo existente para inserir a data na coluna J
        book = openpyxl.load_workbook(file_path_lmc)
        sheet = book.active
        startrow = sheet.max_row

        # Inserir a data copiada na coluna J na última linha ao lado dos dados preenchidos
        sheet.cell(row=startrow, column=10, value=date_str)
        
        # Salvar o arquivo
        book.save(file_path_lmc)
        
        # Verificar se a data é menor ou igual a dois dias atrás ou se a data foi digitada como o primeiro dia do mês
        if not is_date_recent(date_str) or date_str == first_day_of_month:
            # Se a data não for recente ou for inválida, pula para o próximo combustível na lista
            continue
           
        # Segue o processamento adicional do combustível aqui...
        click_position(579, 619)  # Gerar LMC
        time.sleep(10)
        click_position(676, 118)  # Conciliação dos estoques
        
        # Passos para o primeiro tanque
        click_position(590, 236)  # Primeiro tanque coluna perdas
        time.sleep(2)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(2)
        copied_content = pyperclip.paste()
        print("Conteúdo copiado do primeiro tanque:", copied_content)
        data_lines = copied_content.strip().split('\n')
        first_tank_data = []
        if len(data_lines) > 1:
            for line in data_lines[1:]:
                data_columns = re.split(r'\t|\s{2,}', line.strip())
                first_tank_data.append(data_columns)
                for index, value in enumerate(data_columns):
                    column_letter = chr(75 + index)
                    sheet[f'{column_letter}{startrow}'] = value
        else:
            print("Nenhum dado encontrado além dos cabeçalhos.")
        book.save(file_path_lmc)
        click_position(763, 237)  # Primeiro tanque coluna fechamento
        pyautogui.press('delete')
        time.sleep(2)
        click_position(694, 369)  # Clica fora
        click_position(590, 236)  # Primeiro tanque coluna perdas
        time.sleep(2)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(2)
        copied_content = pyperclip.paste()
        print("Conteúdo copiado do primeiro tanque:", copied_content)
        data_lines = copied_content.strip().split('\n')
        if len(data_lines) > 1:
            for line in data_lines[1:]:
                data_columns = re.split(r'\t|\s{2,}', line.strip())
                if len(data_columns) > 1:
                    perdas_value = data_columns[1]
                    try:
                        cleaned_value = perdas_value.replace(' ', '').replace('.', '').replace(',', '.')
                        original_value = float(cleaned_value)
                        if original_value != 0:
                            modified_value = int(original_value * 10) // 10 - 1
                        else:
                            modified_value = original_value
                        formatted_value = f"{modified_value:,}".replace(',', '.')
                        sheet[f'O{startrow}'] = formatted_value
                        click_position(765, 238)  # Primeiro tanque coluna perdas
                        time.sleep(2)
                        pyperclip.copy(formatted_value)
                        pyautogui.hotkey('ctrl', 'v')
                        click_position(690, 372)  # Clica fora
                        book.save(file_path_lmc)
                    except ValueError:
                        print("Erro ao converter o valor:", perdas_value)
        else:
            print("Nenhum dado encontrado além dos cabeçalhos.")
        book.save(file_path_lmc)

        # Validação para o segundo tanque
        click_position(585, 258)  # Segundo tanque coluna perdas
        time.sleep(2)
        text = extract_text_from_region(443, 240, 440, 35)
        print(f"Texto extraído da região: '{text}'")
        
        if any(char.isdigit() for char in text):
            print("Dados encontrados no segundo tanque.")
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(2)
            copied_content = pyperclip.paste()
            print("Conteúdo copiado do segundo tanque:", copied_content)
            data_lines = copied_content.strip().split('\n')
            second_tank_data = []
            if len(data_lines) > 1:
                for line in data_lines[1:]:
                    data_columns = re.split(r'\t|\s{2,}', line.strip())
                    second_tank_data.append(data_columns)
                    for index, value in enumerate(data_columns):
                        column_letter = chr(75 + index)
                        sheet[f'{column_letter}{startrow + 1}'] = value
                        
                        # Repetir as colunas de A até J na linha do segundo tanque
                for col in range(1, 11):
                    sheet.cell(row=startrow + 1, column=col, value=sheet.cell(row=startrow, column=col).value)
            else:
                print("Nenhum dado encontrado além dos cabeçalhos.")
            book.save(file_path_lmc)
            
            click_position(763, 258)  # Segundo tanque coluna perdas aqui é pra ser fechamento
            pyautogui.press('delete')
            time.sleep(2)
            click_position(692, 371)  # Clica fora
            click_position(585, 258)  # Segundo tanque coluna perdas
            time.sleep(2)
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(2)
            copied_content = pyperclip.paste()
            print("Conteúdo copiado do segundo tanque:", copied_content)
            data_lines = copied_content.strip().split('\n')
            if len(data_lines) > 1:
                for line in data_lines[1:]:
                    data_columns = re.split(r'\t|\s{2,}', line.strip())
                    if len(data_columns) > 1:
                        perdas_value = data_columns[1]
                        try:
                            cleaned_value = perdas_value.replace(' ', '').replace('.', '').replace(',', '.')
                            original_value = float(cleaned_value)
                            if original_value != 0:
                                modified_value = int(original_value * 10) // 10 - 1
                            else:
                                modified_value = original_value
                            formatted_value = f"{modified_value:,}".replace(',', '.')
                            sheet[f'O{startrow + 1}'] = formatted_value
                            click_position(765, 258)  # Segundo tanque coluna perdas aqui é pra ser fechamento
                            time.sleep(2)
                            pyperclip.copy(formatted_value)
                            pyautogui.hotkey('ctrl', 'v')
                            click_position(692, 372)  # Clica fora
                            book.save(file_path_lmc)
                        except ValueError:
                            print("Erro ao converter o valor:", perdas_value)
            else:
                print("Nenhum dado encontrado além dos cabeçalhos.")
            book.save(file_path_lmc)
        else:
            print("Nenhum dado encontrado no segundo tanque.")

        # Validação para o terceiro tanque
        click_position(582, 276)  # Terceiro tanque coluna perdas
        time.sleep(2)
        text = extract_text_from_region(443, 268, 440, 35)
        print(f"Texto extraído da região: '{text}'")
        
        if any(char.isdigit() for char in text):
            print("Dados encontrados no terceiro tanque.")
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(2)
            copied_content = pyperclip.paste()
            print("Conteúdo copiado do terceiro tanque:", copied_content)
            data_lines = copied_content.strip().split('\n')
            third_tank_data = []
            if len(data_lines) > 1:
                for line in data_lines[1:]:
                    data_columns = re.split(r'\t|\s{2,}', line.strip())
                    third_tank_data.append(data_columns)
                    for index, value in enumerate(data_columns):
                        column_letter = chr(75 + index)
                        sheet[f'{column_letter}{startrow + 2}'] = value
                        
                        # Repetir as colunas de A até J na linha do terceiro tanque
                for col in range(1, 11):
                    sheet.cell(row=startrow + 2, column=col, value=sheet.cell(row=startrow, column=col).value)
            else:
                print("Nenhum dado encontrado além dos cabeçalhos.")
            book.save(file_path_lmc)
            
            click_position(766, 276)  # Terceiro tanque coluna perdas aqui é pra ser fechamento
            pyautogui.press('delete')
            time.sleep(2)
            click_position(692, 371)  # Clica fora
            click_position(582, 276)  # Terceiro tanque coluna perdas
            time.sleep(1)
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(1)
            copied_content = pyperclip.paste()
            print("Conteúdo copiado do terceiro tanque:", copied_content)
            data_lines = copied_content.strip().split('\n')
            if len(data_lines) > 1:
                for line in data_lines[1:]:
                    data_columns = re.split(r'\t|\s{2,}', line.strip())
                    if len(data_columns) > 1:
                        perdas_value = data_columns[1]
                        try:
                            cleaned_value = perdas_value.replace(' ', '').replace('.', '').replace(',', '.')
                            original_value = float(cleaned_value)
                            if original_value != 0:
                                modified_value = int(original_value * 10) // 10 - 1
                            else:
                                modified_value = original_value
                            formatted_value = f"{modified_value:,}".replace(',', '.')
                            sheet[f'O{startrow + 2}'] = formatted_value
                            click_position(766, 276)  # Terceiro tanque coluna perdas aqui é pra ser fechamento
                            time.sleep(1)
                            pyperclip.copy(formatted_value)
                            pyautogui.hotkey('ctrl', 'v')
                            click_position(692, 372)  # Clica fora
                            book.save(file_path_lmc)
                        except ValueError:
                            print("Erro ao converter o valor:", perdas_value)
            else:
                print("Nenhum dado encontrado além dos cabeçalhos.")
            book.save(file_path_lmc)
        time.sleep(1)
        click_position(680,620)#GRAVAR LMC
        time.sleep(10)
        click_position(682,402)#Pagina gerada com sucesso
        time.sleep(0)
        click_position(770,447)#Deseja imprimir a pagina gerada (NÃO)
        

        # Passos finais para fechar o processo
        click_position(921, 94)  # Fecha a geração do LMC
        

        
# Iterar sobre as linhas da planilha
 for index, row in df_agosto.iterrows():
    if check_stop():
        break
    
    impedimento = row[col_impedimento]
    loja_numero = row[col_loja_numero]

    if impedimento == 'NÃO':
        # Clicar no botão 'Selecionar empresa'
        pyautogui.click(x=169, y=88)
        time.sleep(1)

        # Digitar o número da loja como inteiro
        pyautogui.write(str(int(loja_numero)))
        time.sleep(2)

        # Pressionar Enter
        pyautogui.press('enter')
        time.sleep(30)

        # Clicar na posição para digitar 'lmc automatico'
        pyautogui.click(x=5, y=173)
        pyautogui.typewrite('lmc automatico', interval=0.1)

        # Adicionar os novos cliques
        time.sleep(1)
        
        # Coordenadas para o duplo clique e abrir o LMC
        x, y = 106, 276
        pyautogui.doubleClick(x=x, y=y)
        
        time.sleep(3)  # tempo necessário para abrir o LMC e extrair dados

        # Executar a função principal para automatizar a extração
        automate_extraction()

        # FECHAR LMC 
        pyautogui.click(x=1159, y=26)
        time.sleep(1)
        # CLICAR PARA APAGAR 'lmc automatico'
        pyautogui.click(250, 174)
        time.sleep(1)

        # Palavra a ser apagada
        word_to_delete = "lmc automatico"

        # Número de caracteres na palavra incluindo espaços
        num_characters = len(word_to_delete)

        # Esperar um tempo para garantir que o cursor esteja na posição correta
        time.sleep(1)

        # Apagar todos os caracteres
        pyautogui.typewrite(['backspace'] * num_characters)

        # Escrever 'ok' na coluna E da mesma linha
        ws_agosto[f'E{index + 2}'] = 'ok'  # +2 para ajustar o índice do DataFrame para a linha do Excel

        # Salvar a planilha após cada iteração
        wb_agosto.save(file_path_agosto)

    # Adicionando delay para não sobrecarregar o sistema
    time.sleep(1)

# Salvar a planilha final após completar o processamento
 wb_agosto.save(file_path_agosto)
